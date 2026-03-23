#!/usr/bin/env python3
"""
FF&E Price Analysis - Excel Generator
Modes: "individual" (per-item tables) and "package" (single comparison table)
Markup zone to the right of the visible table, same-row aligned.

Usage:
  python build_comparison.py --config config.json
"""

import argparse, json, csv, sys, os, re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl"); sys.exit(1)

try:
    import pandas as pd
    USE_PD = True
except ImportError:
    USE_PD = False

PROVINCE_TAX = {
    "AB": 5.0, "BC": 12.0, "MB": 12.0, "NB": 15.0, "NL": 15.0,
    "NS": 15.0, "NT": 5.0, "NU": 5.0, "ON": 13.0, "PE": 15.0,
    "QC": 14.975, "SK": 11.0, "YT": 5.0,
}

# Service descriptions that should NOT get markup (normalized for matching)
SERVICE_TERMS = {"freight", "installation", "install", "removal", "remove",
                 "setinplace", "set in place", "sorac", "ecofrais", "delivery",
                 # French equivalents (ASCII-folded; accents stripped before matching)
                 "livraison", "fret", "frais de transport", "fraisdetransport",
                 "mise en place", "miseenplace", "enlevement"}

# Accent folding map: accented char -> ASCII equivalent
_ACCENT_MAP = str.maketrans(
    "àâäáãåéèêëíîïìóôöòõúûüùçñ"
    "ÀÂÄÁÃÅÉÈÊËÍÎÏÌÓÔÖÒÕÚÛÜÙÇÑ",
    "aaaaaaeeeeiiiiooooouuuucn"
    "AAAAAAEEEEIIIIOOOOOUUUUCN")

def _fold_accents(s):
    """Replace accented characters with their ASCII base letter."""
    return s.translate(_ACCENT_MAP)


# ===================== I18N LABELS =====================

_LABELS = {
    "en": {
        "sheet_name":           "Price Analysis",
        "default_title_ind":    "Equipment Price Comparison",
        "default_title_pkg":    "Package Price Comparison",
        "details":              "DETAILS",
        "image":                "IMAGE",
        "incl_in_total":        "Incl. in Total",
        "qty":                  "QTY",
        "line_item":            "LINE ITEM",
        "brand":                "BRAND",
        "code":                 "CODE",
        "description":          "DESCRIPTION",
        "brand_label":          "BRAND",
        "code_label":           "CODE",
        "power_label":          "POWER",
        "total_before_tax":     "Total Before Tax",
        "tax":                  "Tax",
        "subtotal":             "Subtotal",
        "total":                "TOTAL",
        "best_price":           ">> BEST PRICE <<",
        "specs_header":         "SPECS & DESCRIPTION",
        "specs_placeholder":    "Enter item specifications, dimensions, and technical details here...",
        "markup_zone":          "MARKUP ZONE",
        "markup_pct":           "MARKUP %",
        "apply":                "APPLY?",
        "included_in_total":    "Included in Total",
    },
    "fr": {
        "sheet_name":           "Analyse de prix",
        "default_title_ind":    "Comparaison de prix — Équipements",
        "default_title_pkg":    "Comparaison de prix — Forfait",
        "details":              "DÉTAILS",
        "image":                "IMAGE",
        "incl_in_total":        "Incl. au total",
        "qty":                  "QTÉ",
        "line_item":            "ARTICLE",
        "brand":                "MARQUE",
        "code":                 "CODE",
        "description":          "DESCRIPTION",
        "brand_label":          "MARQUE",
        "code_label":           "CODE",
        "power_label":          "ALIMENTATION",
        "total_before_tax":     "Total avant taxe",
        "tax":                  "Taxe",
        "subtotal":             "Sous-total",
        "total":                "TOTAL",
        "best_price":           ">> MEILLEUR PRIX <<",
        "specs_header":         "SPÉCIFICATIONS ET DESCRIPTION",
        "specs_placeholder":    "Inscrire les spécifications, dimensions et détails techniques ici...",
        "markup_zone":          "ZONE MAJORATION",
        "markup_pct":           "MAJORATION %",
        "apply":                "APPLIQUER?",
        "included_in_total":    "Inclus au total",
    },
}

def L(key, cfg):
    """Return the localized label for the given key. Falls back to English."""
    lang = cfg.get("lang", "en").lower()
    if lang not in _LABELS:
        lang = "en"
    return _LABELS[lang].get(key, _LABELS["en"].get(key, key))


def is_service(desc):
    """Check if a description is a known service term (no markup).
    Handles French accented characters (É→E, è→e, etc.)."""
    folded = _fold_accents(desc.lower().strip())
    normalized = re.sub(r'[^a-z0-9 ]', '', folded)
    normalized_nospace = normalized.replace(" ", "")
    return normalized in SERVICE_TERMS or normalized_nospace in SERVICE_TERMS

def apply_exclude_cf(ws, inc_col_letter, sdr, edr):
    """Apply conditional formatting: ✗ cells turn red font."""
    red_font = Font(name='Segoe UI', bold=True, color="E74C3C")
    dxf = DifferentialStyle(font=red_font)
    rule = Rule(type="containsText", operator="containsText", text="✗", dxf=dxf)
    rule.formula = [f'NOT(ISERROR(SEARCH("✗",{inc_col_letter}{sdr})))']
    range_str = f"{inc_col_letter}{sdr}:{inc_col_letter}{edr}"
    ws.conditional_formatting.add(range_str, rule)

# -- Styles --
HEADER_BLUE = PatternFill("solid", fgColor="288AD6")
DETAILS_BG  = PatternFill("solid", fgColor="FAFAFA")
COL_HDR_BG  = PatternFill("solid", fgColor="F8F9FA")
WINNER_BG   = PatternFill("solid", fgColor="F2FAF2")
SPECS_BG    = PatternFill("solid", fgColor="FAFAFA")
SEP_FILL    = PatternFill("solid", fgColor="E0E7EF")
WHITE_BG    = PatternFill("solid", fgColor="FFFFFF")
ALT_ROW_BG  = PatternFill("solid", fgColor="F5F7FA")
MARKUP_HDR  = PatternFill("solid", fgColor="FFF3CD")  # warm yellow for markup header
MARKUP_BG   = PatternFill("solid", fgColor="FFFDF0")  # very light yellow

SUBTLE_BD  = Border(bottom=Side(style='thin', color="F0F0F0"))
COL_HDR_BD = Border(bottom=Side(style='medium', color="E5E5E5"))
TOTAL_BD   = Border(top=Side(style='medium', color="288AD6"))
SEP_BD     = Border(bottom=Side(style='medium', color="288AD6"))
THIN_BD    = Border(
    left=Side(style='thin', color="E5E5E5"), right=Side(style='thin', color="E5E5E5"),
    top=Side(style='thin', color="E5E5E5"), bottom=Side(style='thin', color="E5E5E5"))

TP = "1D1D1F"; TS = "86868B"; WH = "FFFFFF"


# ===================== DATA HELPERS =====================

def load_csv(path):
    if USE_PD:
        df = pd.read_csv(path, keep_default_na=False)
        for c in ["type","supplier","brand","code","description","Power Type"]:
            if c in df.columns: df[c] = df[c].astype(str).str.strip()
        df["price"] = pd.to_numeric(df["price"], errors="coerce").fillna(0)
        return df
    else:
        rows = []
        with open(path, newline='', encoding='utf-8-sig') as f:
            rd = csv.DictReader(f); rd.fieldnames = [h.strip() for h in rd.fieldnames]
            for r in rd:
                row = {k.strip(): v.strip() for k, v in r.items()}
                try: row["price"] = float(row.get("price","0").replace("$","").replace(",",""))
                except ValueError: row["price"] = 0
                rows.append(row)
        return rows

def groups_from(df):
    if USE_PD:
        mi = df[(df["type"]=="item") & df["Power Type"].notna() & (df["Power Type"]!="") & (df["Power Type"]!="nan")]
        seen, out = set(), []
        for _, r in mi.iterrows():
            k = (r["code"], r["Power Type"])
            if k not in seen: seen.add(k); out.append(k)
        return out
    else:
        seen, out = set(), []
        for r in df:
            if r.get("type")=="item" and r.get("Power Type",""):
                k = (r["code"], r["Power Type"])
                if k not in seen: seen.add(k); out.append(k)
        return out

def items_for(df, code, pt):
    if USE_PD:
        return df[(df["code"]==code) & ((df["Power Type"]==pt)|(df["Power Type"].isna())|(df["Power Type"]=="")|(df["Power Type"]=="nan")) & (df["type"].isin(["item","subitem"]))]
    else:
        return [r for r in df if r["code"]==code and r.get("type") in ("item","subitem") and (r.get("Power Type","")==pt or not r.get("Power Type",""))]

def uniq_col(items, col):
    if USE_PD: return list(items[col].unique())
    else:
        seen, out = [], []
        for r in items:
            v = r.get(col,"")
            if v not in seen: seen.append(v); out.append(v)
        return out

def first_brand(items):
    if USE_PD:
        mi = items[items["type"]=="item"]
        return mi.iloc[0]["brand"] if not mi.empty else ""
    else:
        for r in items:
            if r["type"]=="item": return r.get("brand","")
        return ""

def price_for(items, sup, desc):
    if USE_PD:
        r = items[(items["supplier"]==sup)&(items["description"]==desc)]
        return float(r["price"].iloc[0]) if not r.empty else 0
    else:
        for r in items:
            if r["supplier"]==sup and r["description"]==desc: return float(r.get("price",0))
        return 0

def all_suppliers(df):
    if USE_PD: return list(df["supplier"].unique())
    else:
        seen, out = [], []
        for r in df:
            s = r.get("supplier","")
            if s and s not in seen: seen.append(s); out.append(s)
        return out

def all_items_flat(df):
    if USE_PD:
        mi = df[df["type"]=="item"]
        seen, out = set(), []
        for _, r in mi.iterrows():
            k = (r["code"], r["description"])
            if k not in seen:
                seen.add(k)
                out.append({"code": r["code"], "description": r["description"],
                            "brand": r["brand"], "power": r["Power Type"]})
        return out
    else:
        seen, out = set(), []
        for r in df:
            if r.get("type")=="item":
                k = (r["code"], r["description"])
                if k not in seen:
                    seen.add(k)
                    out.append({"code": r["code"], "description": r["description"],
                                "brand": r.get("brand",""), "power": r.get("Power Type","")})
        return out

def price_for_flat(df, sup, code, desc):
    if USE_PD:
        r = df[(df["supplier"]==sup)&(df["code"]==code)&(df["description"]==desc)&(df["type"]=="item")]
        return float(r["price"].iloc[0]) if not r.empty else 0
    else:
        for r in df:
            if r["supplier"]==sup and r["code"]==code and r["description"]==desc and r["type"]=="item":
                return float(r.get("price",0))
        return 0


# ===================== MARKUP ZONE HELPERS =====================

def get_markup_pct(cfg, supplier):
    """Get markup % for a supplier from config. Default 0."""
    m = cfg.get("markup", {})
    return m.get(supplier, 0)

def write_markup_header(ws, row, mk_start, sups, cfg):
    """Write the markup zone header: title + per-supplier markup % cells."""
    # Title
    c = ws.cell(row=row, column=mk_start, value=L("markup_zone", cfg))
    c.font = Font(name='Segoe UI', bold=True, size=10, color=TP)
    c.fill = MARKUP_HDR

    # Per-supplier: 2 columns each (Base, Apply?)
    for si, s in enumerate(sups):
        base_col = mk_start + 1 + si * 2
        apply_col = base_col + 1
        # Supplier name spanning both columns
        ws.merge_cells(start_row=row, start_column=base_col, end_row=row, end_column=apply_col)
        c = ws.cell(row=row, column=base_col, value=s)
        c.font = Font(name='Segoe UI', bold=True, size=9, color=TS)
        c.fill = MARKUP_HDR
        c.alignment = Alignment(horizontal="center")

    # Markup % row (row+1)
    pct_row = row + 1
    c = ws.cell(row=pct_row, column=mk_start, value=L("markup_pct", cfg))
    c.font = Font(name='Segoe UI', bold=True, size=9, color=TS)
    c.fill = MARKUP_HDR

    pct_cells = {}
    for si, s in enumerate(sups):
        base_col = mk_start + 1 + si * 2
        pct = get_markup_pct(cfg, s)
        c = ws.cell(row=pct_row, column=base_col, value=pct / 100.0 if pct else 0)
        c.number_format = '0.00%'
        c.font = Font(name='Segoe UI', bold=True, size=11, color="D35400")
        c.fill = MARKUP_HDR
        c.alignment = Alignment(horizontal="center")
        pct_cells[s] = f"${get_column_letter(base_col)}${pct_row}"

        # Label for Apply column
        c2 = ws.cell(row=pct_row, column=base_col + 1, value=L("apply", cfg))
        c2.font = Font(name='Segoe UI', bold=True, size=9, color=TS)
        c2.fill = MARKUP_HDR
        c2.alignment = Alignment(horizontal="center")

    return pct_cells  # dict of supplier -> absolute cell ref for their markup %

def write_markup_row(ws, row, mk_start, sups, pct_cells, desc, prices, dv):
    """Write one markup zone row: base price + YES/NO per supplier."""
    # Description echo
    c = ws.cell(row=row, column=mk_start, value=desc)
    c.font = Font(name='Segoe UI', size=9, color=TS)
    c.fill = MARKUP_BG

    default_apply = "NO" if is_service(desc) else "YES"

    apply_refs = {}
    for si, s in enumerate(sups):
        base_col = mk_start + 1 + si * 2
        apply_col = base_col + 1

        base_price = prices.get(s, 0)

        # Base price
        c = ws.cell(row=row, column=base_col, value=base_price if base_price > 0 else "")
        if base_price > 0:
            c.number_format = '$#,##0.00'
        c.font = Font(name='Segoe UI', size=9, color=TS)
        c.fill = MARKUP_BG
        c.alignment = Alignment(horizontal="right")

        # Apply? YES/NO
        c2 = ws.cell(row=row, column=apply_col, value=default_apply)
        c2.font = Font(name='Segoe UI', size=9, color=TP, bold=True)
        c2.fill = MARKUP_BG
        c2.alignment = Alignment(horizontal="center")
        dv.add(c2)

        apply_refs[s] = {
            "base_cell": f"{get_column_letter(base_col)}{row}",
            "apply_cell": f"{get_column_letter(apply_col)}{row}",
            "pct_cell": pct_cells[s],
        }

    return apply_refs

def markup_formula(qty_cell, refs):
    """Build the visible table formula. Prices always visible; SUMIF handles exclusion."""
    bc = refs["base_cell"]
    ac = refs["apply_cell"]
    pc = refs["pct_cell"]
    return f'=IF({bc}="","",IF({ac}="YES",{qty_cell}*{bc}*(1+{pc}),{qty_cell}*{bc}))'


# ===================== INDIVIDUAL MODE =====================

def build_individual(df, cfg):
    tr = cfg["tax_rate"] / 100.0
    tl = cfg.get("tax_label", f"{cfg['tax_rate']}%")
    title = cfg.get("title", L("default_title_ind", cfg))
    specs = cfg.get("include_specs", False)
    show_toggle = cfg.get("include_toggle", True)
    wb = Workbook(); ws = wb.active
    ws.title = L("sheet_name", cfg); ws.sheet_view.showGridLines = False

    # YES/NO data validation
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.error = "Use YES or NO"
    ws.add_data_validation(dv)

    groups = groups_from(df)
    if not groups: print("ERROR: No product groups found"); sys.exit(1)

    cr = 1
    # Track max supplier count for markup zone positioning
    max_sups = max(len(uniq_col(items_for(df, c, p), "supplier")) for c, p in groups)
    # Column layout: A(margin) + B(details) + C(image) + [D(inc)] + E(qty) + F(desc) + suppliers + gap
    # image column always present in individual mode
    if show_toggle:
        COL_IMG_IND = 3; COL_INC_BASE = 4; COL_QTY_BASE = 5; COL_DESC_BASE = 6; SUP_START_BASE = 7
    else:
        COL_IMG_IND = 3; COL_INC_BASE = None; COL_QTY_BASE = 4; COL_DESC_BASE = 5; SUP_START_BASE = 6
    MK_START = SUP_START_BASE + max_sups + 2  # 2 column gap

    # Collect ALL unique suppliers across all groups for the markup zone header
    all_sups = []
    for c2, p2 in groups:
        for s in uniq_col(items_for(df, c2, p2), "supplier"):
            if s not in all_sups: all_sups.append(s)

    # Write markup header on rows 1-2 (will align with title)
    pct_cells = write_markup_header(ws, 1, MK_START, all_sups, cfg)

    for gi, (code, pt) in enumerate(groups, 1):
        items = items_for(df, code, pt)
        sups = uniq_col(items, "supplier")
        brand = first_brand(items)
        descs = uniq_col(items, "description")
        last_col = SUP_START_BASE + len(sups) - 1  # dynamic per group, same logic as original

        # Winner (uses marked-up prices for fair comparison)
        winner = ""
        mnt = float('inf')
        for s in sups:
            mkp = get_markup_pct(cfg, s) / 100.0
            t = 0
            for d in descs:
                p = price_for(items, s, d)
                if p > 0:
                    t += p * (1 + mkp) if not is_service(d) else p
            if t > 0 and t * (1 + tr) < mnt:
                mnt = t * (1 + tr); winner = s

        # -- Header --
        if gi == 1:
            ws.row_dimensions[cr].height = 40
            ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=last_col)
            c = ws.cell(row=cr, column=2, value=title)
            c.font = Font(name='Segoe UI', bold=True, size=14, color=WH)
            c.fill = HEADER_BLUE
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            cr += 2  # skip title + markup% row (markup header already written)
        else:
            ws.row_dimensions[cr].height = 4
            for ci in range(2, last_col + 1):
                sc = ws.cell(row=cr, column=ci); sc.fill = SEP_FILL; sc.border = SEP_BD
            cr += 1

        # -- Column headers -- cols: 2=DETAILS,3=IMAGE,[4=INC,]5=QTY,6=LINE ITEM,7+=sups
        ws.row_dimensions[cr].height = 32
        _incl = L("incl_in_total", cfg)
        hdr_ind = [L("details", cfg), L("image", cfg)] + ([_incl] if show_toggle else []) + [L("qty", cfg), L("line_item", cfg)] + sups
        for i, h in enumerate(hdr_ind):
            ci = i + 2
            c = ws.cell(row=cr, column=ci, value=h if h in ("✓", _incl) else h.upper())
            c.font = Font(name='Segoe UI', size=9, bold=True, color=TS)
            c.fill = WINNER_BG if h == winner else COL_HDR_BG
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = COL_HDR_BD
        cr += 1

        sdr = cr; ni = len(descs)

        # Details merged
        dc = ws.cell(row=sdr, column=2, value=f"{L('brand_label', cfg)}\n{brand}\n\n{L('code_label', cfg)}\n{code}\n\n{L('power_label', cfg)}\n{pt}")
        dc.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=1)
        dc.font = Font(name='Segoe UI', size=10, color=TS); dc.fill = DETAILS_BG
        me = sdr + ni + 1
        ws.merge_cells(start_row=sdr, start_column=2, end_row=me, end_column=2)

        # Image: always present — individual white cells
        for ir in range(sdr, me + 2):
            ic = ws.cell(row=ir, column=COL_IMG_IND, value=""); ic.fill = WHITE_BG; ic.border = SUBTLE_BD

        # Include DV for this group (only if toggle enabled)
        if show_toggle:
            inc_dv_ind = DataValidation(type="list", formula1='"✓,✗"', allow_blank=False)
            ws.add_data_validation(inc_dv_ind)

        # Item rows — use named column constants
        COL_INC_IND = COL_INC_BASE; COL_QTY_IND = COL_QTY_BASE; COL_DESC_IND = COL_DESC_BASE; SUP_START_IND = SUP_START_BASE
        ql = get_column_letter(COL_QTY_IND)
        for idx, desc in enumerate(descs):
            r = sdr + idx; ws.row_dimensions[r].height = 32

            # Include toggle (optional)
            if show_toggle:
                inc_c = ws.cell(row=r, column=COL_INC_IND, value="✓")
                inc_c.font = Font(name='Segoe UI', size=11, color="2C3E50", bold=True)
                inc_c.alignment = Alignment(horizontal="center", vertical="center")
                inc_c.border = SUBTLE_BD
                inc_dv_ind.add(inc_c)
            inc_cell = f"{get_column_letter(COL_INC_IND)}{r}" if show_toggle else None

            ws.cell(row=r, column=COL_QTY_IND, value=1).font = Font(name='Segoe UI', size=11, color=TP, bold=True)
            ws.cell(row=r, column=COL_QTY_IND).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=COL_QTY_IND).border = SUBTLE_BD

            ws.cell(row=r, column=COL_DESC_IND, value=desc).font = Font(name='Segoe UI', size=11, color=TP)
            ws.cell(row=r, column=COL_DESC_IND).alignment = Alignment(vertical="center", horizontal="left")
            ws.cell(row=r, column=COL_DESC_IND).border = SUBTLE_BD

            # Markup zone row (uses all_sups for consistent column alignment)
            all_prices = {s: price_for(items, s, desc) for s in all_sups}
            refs = write_markup_row(ws, r, MK_START, all_sups, pct_cells, desc, all_prices, dv)

            # Visible table: formula references markup zone (only per-group suppliers)
            qty_cell = f"{ql}{r}"
            for si, s in enumerate(sups):
                col = SUP_START_IND + si
                p = all_prices[s]
                c = ws.cell(row=r, column=col)
                if p > 0:
                    c.value = markup_formula(qty_cell, refs[s])
                    c.number_format = '$#,##0.00'
                    c.font = Font(name='Segoe UI', size=11, color=TP)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.value = ""
                c.border = SUBTLE_BD
                if s == winner: c.fill = WINNER_BG

        # Apply red font CF for ✗ rows (only if toggle enabled)
        if show_toggle:
            apply_exclude_cf(ws, get_column_letter(COL_INC_IND), sdr, sdr + ni - 1)

        # Total Before Tax
        tbr = sdr + ni; ws.row_dimensions[tbr].height = 32
        merge_start = COL_INC_IND if show_toggle else COL_QTY_IND
        ws.merge_cells(start_row=tbr, start_column=merge_start, end_row=tbr, end_column=COL_DESC_IND)
        lc = ws.cell(row=tbr, column=merge_start, value=L("total_before_tax", cfg))
        lc.font = Font(name='Segoe UI', size=11, bold=True, color=TP)
        lc.alignment = Alignment(vertical="center", horizontal="left"); lc.border = SUBTLE_BD
        for si, s in enumerate(sups):
            col = SUP_START_IND + si; cl = get_column_letter(col)
            if show_toggle:
                inc_cl = get_column_letter(COL_INC_IND)
                formula = f'=SUMIF({inc_cl}{sdr}:{inc_cl}{tbr-1},"✓",{cl}{sdr}:{cl}{tbr-1})'
            else:
                formula = f'=SUM({cl}{sdr}:{cl}{tbr-1})'
            c = ws.cell(row=tbr, column=col, value=formula)
            c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
            c.font = Font(name='Segoe UI', size=11, bold=True, color=TP); c.border = SUBTLE_BD
            if s == winner: c.fill = WINNER_BG

        # Tax
        txr = tbr + 1; ws.row_dimensions[txr].height = 28
        ws.merge_cells(start_row=txr, start_column=merge_start, end_row=txr, end_column=COL_DESC_IND)
        tc = ws.cell(row=txr, column=merge_start, value=f"{L('tax', cfg)} {tl}")
        tc.font = Font(name='Segoe UI', size=10, color=TS)
        tc.alignment = Alignment(vertical="center", horizontal="left"); tc.border = SUBTLE_BD
        for si, s in enumerate(sups):
            col = SUP_START_IND + si; cl = get_column_letter(col)
            c = ws.cell(row=txr, column=col, value=f"={cl}{tbr}*{tr}")
            c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
            c.font = Font(name='Segoe UI', size=10, color=TS); c.border = SUBTLE_BD
            if s == winner: c.fill = WINNER_BG

        # Final Total
        ttr = txr + 1; ws.row_dimensions[ttr].height = 40
        ws.merge_cells(start_row=ttr, start_column=2, end_row=ttr, end_column=COL_DESC_IND)
        ws.cell(row=ttr, column=2, value="").fill = DETAILS_BG
        for si, s in enumerate(sups):
            col = SUP_START_IND + si; cl = get_column_letter(col)
            c = ws.cell(row=ttr, column=col, value=f"={cl}{tbr}+{cl}{txr}")
            c.font = Font(name='Segoe UI', bold=True, size=13, color=TP)
            c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = TOTAL_BD
            if s == winner: c.fill = WINNER_BG

        cr = ttr + 1

        if specs:
            ws.row_dimensions[cr].height = 60
            ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=last_col)
            sc = ws.cell(row=cr, column=2,
                value=f"{L('specs_header', cfg)}\n\n{L('specs_placeholder', cfg)}")
            sc.font = Font(name='Segoe UI', size=10, color=TS)
            sc.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=2)
            sc.fill = SPECS_BG; sc.border = Border(bottom=Side(style='thin', color="F0F0F0"))
            cr += 1
        cr += 2

    # Column widths
    ws.column_dimensions['A'].width = 2; ws.column_dimensions['B'].width = 18
    ws.column_dimensions[get_column_letter(COL_IMG_IND)].width = 12  # IMAGE always present
    if show_toggle:
        ws.column_dimensions[get_column_letter(COL_INC_BASE)].width = 12   # INC
        ws.column_dimensions[get_column_letter(COL_QTY_BASE)].width = 7
        ws.column_dimensions[get_column_letter(COL_DESC_BASE)].width = 35
    else:
        ws.column_dimensions[get_column_letter(COL_QTY_BASE)].width = 7
        ws.column_dimensions[get_column_letter(COL_DESC_BASE)].width = 35
    for i in range(max_sups): ws.column_dimensions[get_column_letter(SUP_START_BASE+i)].width = 16
    # Markup zone widths
    ws.column_dimensions[get_column_letter(MK_START)].width = 20
    for si in range(len(all_sups)):
        ws.column_dimensions[get_column_letter(MK_START + 1 + si*2)].width = 14
        ws.column_dimensions[get_column_letter(MK_START + 2 + si*2)].width = 8

    wb.save(cfg["output_xlsx"])
    print(f"[individual] Saved: {cfg['output_xlsx']} | {len(groups)} groups | Tax: {tl}")
    mk = cfg.get("markup", {})
    if any(v > 0 for v in mk.values()):
        print(f"Markup: {mk}")


# ===================== PACKAGE MODE =====================

def build_package(df, cfg):
    tr = cfg["tax_rate"] / 100.0
    tl = cfg.get("tax_label", f"{cfg['tax_rate']}%")
    title = cfg.get("title", L("default_title_pkg", cfg))
    show_img = cfg.get("include_image", False)

    wb = Workbook(); ws = wb.active
    ws.title = L("sheet_name", cfg); ws.sheet_view.showGridLines = False

    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.error = "Use YES or NO"; ws.add_data_validation(dv)

    sups = all_suppliers(df)
    items = all_items_flat(df)
    if not items: print("ERROR: No items found"); sys.exit(1)

    if show_img:
        COL_BRAND=2; COL_CODE=3; COL_DESC=4; COL_IMG=5; COL_QTY=6; SUP_START=7
    else:
        COL_BRAND=2; COL_CODE=3; COL_DESC=4; COL_IMG=None; COL_QTY=5; SUP_START=6

    last_col = SUP_START + len(sups) - 1
    MK_START = last_col + 3  # 2 column gap

    cr = 1

    # Title
    ws.row_dimensions[cr].height = 40
    ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=last_col)
    tc = ws.cell(row=cr, column=2, value=title)
    tc.font = Font(name='Segoe UI', bold=True, size=14, color=WH)
    tc.fill = HEADER_BLUE
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # Markup header (same rows as title)
    pct_cells = write_markup_header(ws, cr, MK_START, sups, cfg)
    cr += 2

    # Winner
    pkg_winner = ""
    pkg_min = float('inf')
    for s in sups:
        mkp = get_markup_pct(cfg, s) / 100.0
        total = 0
        for it in items:
            p = price_for_flat(df, s, it["code"], it["description"])
            if p > 0:
                total += p * (1 + mkp) if not is_service(it["description"]) else p
        if total > 0 and total * (1 + tr) < pkg_min:
            pkg_min = total * (1 + tr); pkg_winner = s

    # Column headers
    ws.row_dimensions[cr].height = 28
    _incl_pkg = L("included_in_total", cfg)
    hdr_labels = [L("brand", cfg), L("code", cfg), L("description", cfg)]
    if show_img: hdr_labels.append(L("image", cfg))
    hdr_labels += [L("qty", cfg)]
    hdr_labels += sups
    for i, h in enumerate(hdr_labels):
        ci = i + 2
        c = ws.cell(row=cr, column=ci, value=h if h in ("✓", _incl_pkg) else h.upper())
        c.font = Font(name='Segoe UI', size=9, bold=True, color=TS)
        c.fill = WINNER_BG if h == pkg_winner else COL_HDR_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = COL_HDR_BD
    cr += 1

    # Item rows
    sdr = cr
    ql = get_column_letter(COL_QTY)

    for idx, it in enumerate(items):
        r = cr; ws.row_dimensions[r].height = 30
        fill = ALT_ROW_BG if idx % 2 == 1 else WHITE_BG

        ws.cell(row=r, column=COL_BRAND, value=it["brand"]).font = Font(name='Segoe UI', size=10, color=TP)
        ws.cell(row=r, column=COL_BRAND).fill = fill; ws.cell(row=r, column=COL_BRAND).border = THIN_BD

        ws.cell(row=r, column=COL_CODE, value=it["code"]).font = Font(name='Segoe UI', size=10, color=TP)
        ws.cell(row=r, column=COL_CODE).fill = fill; ws.cell(row=r, column=COL_CODE).border = THIN_BD

        ws.cell(row=r, column=COL_DESC, value=it["description"]).font = Font(name='Segoe UI', size=10, color=TP)
        ws.cell(row=r, column=COL_DESC).alignment = Alignment(vertical="center", horizontal="left")
        ws.cell(row=r, column=COL_DESC).fill = fill; ws.cell(row=r, column=COL_DESC).border = THIN_BD

        if show_img:
            ws.cell(row=r, column=COL_IMG, value="").fill = WHITE_BG
            ws.cell(row=r, column=COL_IMG).border = THIN_BD

        ws.cell(row=r, column=COL_QTY, value=1).font = Font(name='Segoe UI', size=10, color=TP, bold=True)
        ws.cell(row=r, column=COL_QTY).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=COL_QTY).fill = fill; ws.cell(row=r, column=COL_QTY).border = THIN_BD

        # Markup zone row
        prices = {s: price_for_flat(df, s, it["code"], it["description"]) for s in sups}
        refs = write_markup_row(ws, r, MK_START, sups, pct_cells, it["description"], prices, dv)

        # Supplier prices (formula from markup zone)
        qty_cell = f"{ql}{r}"
        for si, s in enumerate(sups):
            col = SUP_START + si
            p = prices[s]
            c = ws.cell(row=r, column=col)
            if p > 0:
                c.value = markup_formula(qty_cell, refs[s])
                c.number_format = '$#,##0.00'
                c.font = Font(name='Segoe UI', size=10, color=TP)
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.value = ""
            c.fill = fill; c.border = THIN_BD
            if s == pkg_winner: c.fill = WINNER_BG

        cr += 1

    # Subtotal
    merge_end = COL_QTY
    ws.row_dimensions[cr].height = 32
    ws.merge_cells(start_row=cr, start_column=COL_BRAND, end_row=cr, end_column=merge_end)

    lc = ws.cell(row=cr, column=COL_BRAND, value=L("subtotal", cfg))
    lc.font = Font(name='Segoe UI', size=11, bold=True, color=TP)
    lc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    lc.border = Border(top=Side(style='medium', color="E5E5E5"))
    for si, s in enumerate(sups):
        col = SUP_START+si; cl = get_column_letter(col)
        c = ws.cell(row=cr, column=col, value=f'=SUM({cl}{sdr}:{cl}{cr-1})')
        c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(name='Segoe UI', size=11, bold=True, color=TP)
        c.border = Border(top=Side(style='medium', color="E5E5E5"))
        if s == pkg_winner: c.fill = WINNER_BG
    sub_row = cr; cr += 1

    # Tax
    ws.row_dimensions[cr].height = 28
    ws.merge_cells(start_row=cr, start_column=COL_BRAND, end_row=cr, end_column=merge_end)
    tc = ws.cell(row=cr, column=COL_BRAND, value=f"{L('tax', cfg)} {tl}")
    tc.font = Font(name='Segoe UI', size=10, color=TS)
    tc.alignment = Alignment(vertical="center", horizontal="right", indent=1); tc.border = SUBTLE_BD
    for si, s in enumerate(sups):
        col = SUP_START+si; cl = get_column_letter(col)
        c = ws.cell(row=cr, column=col, value=f"={cl}{sub_row}*{tr}")
        c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(name='Segoe UI', size=10, color=TS); c.border = SUBTLE_BD
        if s == pkg_winner: c.fill = WINNER_BG
    tax_r = cr; cr += 1

    # Total
    ws.row_dimensions[cr].height = 40
    ws.merge_cells(start_row=cr, start_column=COL_BRAND, end_row=cr, end_column=merge_end)
    tl_cell = ws.cell(row=cr, column=COL_BRAND, value=L("total", cfg))
    tl_cell.font = Font(name='Segoe UI', size=13, bold=True, color=TP)
    tl_cell.alignment = Alignment(vertical="center", horizontal="right", indent=1); tl_cell.border = TOTAL_BD
    for si, s in enumerate(sups):
        col = SUP_START+si; cl = get_column_letter(col)
        c = ws.cell(row=cr, column=col, value=f"={cl}{sub_row}+{cl}{tax_r}")
        c.font = Font(name='Segoe UI', bold=True, size=13, color=TP)
        c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = TOTAL_BD
        if s == pkg_winner: c.fill = WINNER_BG
    cr += 1

    # Winner row
    ws.row_dimensions[cr].height = 28
    ws.merge_cells(start_row=cr, start_column=COL_BRAND, end_row=cr, end_column=merge_end)
    ws.cell(row=cr, column=COL_BRAND, value="").fill = WINNER_BG
    for si, s in enumerate(sups):
        col = SUP_START + si; c = ws.cell(row=cr, column=col)
        if s == pkg_winner:
            c.value = L("best_price", cfg); c.font = Font(name='Segoe UI', bold=True, size=11, color="1B5E20")
        else: c.value = ""
        c.fill = WINNER_BG; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style='thin', color="E5E5E5"))

    # Column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions[get_column_letter(COL_BRAND)].width = 14
    ws.column_dimensions[get_column_letter(COL_CODE)].width = 16
    ws.column_dimensions[get_column_letter(COL_DESC)].width = 32
    if show_img: ws.column_dimensions[get_column_letter(COL_IMG)].width = 10
    ws.column_dimensions[get_column_letter(COL_QTY)].width = 7
    for i in range(len(sups)): ws.column_dimensions[get_column_letter(SUP_START+i)].width = 16
    # Markup zone widths
    ws.column_dimensions[get_column_letter(MK_START)].width = 20
    for si in range(len(sups)):
        ws.column_dimensions[get_column_letter(MK_START + 1 + si*2)].width = 14
        ws.column_dimensions[get_column_letter(MK_START + 2 + si*2)].width = 8

    wb.save(cfg["output_xlsx"])
    print(f"[package] Saved: {cfg['output_xlsx']} | {len(items)} items x {len(sups)} suppliers | Tax: {tl}")
    mk = cfg.get("markup", {})
    if any(v > 0 for v in mk.values()): print(f"Markup: {mk}")


# ===================== PACKAGE DETAIL MODE =====================

def all_items_with_subitems(df):
    """Return ordered list of (type, code, description, brand, power) preserving item->subitem structure."""
    if USE_PD:
        seen_items = set()
        out = []
        # Get ordered unique items preserving first-seen order
        for _, r in df.iterrows():
            if r["type"] == "item":
                k = (r["code"], r["description"])
                if k not in seen_items:
                    seen_items.add(k)
                    out.append({"type": "item", "code": r["code"], "description": r["description"],
                                "brand": r["brand"], "power": r["Power Type"]})
                    # Now find all subitems for this code, in order
                    subs = df[(df["type"] == "subitem") & (df["code"] == r["code"])]
                    seen_subs = set()
                    for _, sr in subs.iterrows():
                        sk = sr["description"]
                        if sk not in seen_subs:
                            seen_subs.add(sk)
                            out.append({"type": "subitem", "code": sr["code"],
                                        "description": sr["description"],
                                        "brand": sr["brand"], "power": sr["Power Type"]})
        return out
    else:
        seen_items = set()
        out = []
        for r in df:
            if r.get("type") == "item":
                k = (r["code"], r["description"])
                if k not in seen_items:
                    seen_items.add(k)
                    out.append({"type": "item", "code": r["code"], "description": r["description"],
                                "brand": r.get("brand",""), "power": r.get("Power Type","")})
                    seen_subs = set()
                    for sr in df:
                        if sr.get("type") == "subitem" and sr.get("code") == r["code"]:
                            sk = sr["description"]
                            if sk not in seen_subs:
                                seen_subs.add(sk)
                                out.append({"type": "subitem", "code": sr["code"],
                                            "description": sr["description"],
                                            "brand": sr.get("brand",""), "power": sr.get("Power Type","")})
        return out

def price_for_any_type(df, sup, code, desc):
    """Get price for either item or subitem."""
    if USE_PD:
        r = df[(df["supplier"]==sup) & (df["code"]==code) & (df["description"]==desc)]
        return float(r["price"].iloc[0]) if not r.empty else 0
    else:
        for r in df:
            if r["supplier"]==sup and r["code"]==code and r["description"]==desc:
                return float(r.get("price", 0))
        return 0

SUBITEM_BG      = PatternFill("solid", fgColor="F0F4FA")
SUBITEM_ALT_BG  = PatternFill("solid", fgColor="E8EEF7")
SUBITEM_WIN_BG  = PatternFill("solid", fgColor="E8F5E9")

def build_package_detail(df, cfg):
    tr = cfg["tax_rate"] / 100.0
    tl = cfg.get("tax_label", f"{cfg['tax_rate']}%")
    title = cfg.get("title", L("default_title_pkg", cfg))
    show_img = cfg.get("include_image", False)
    show_toggle = cfg.get("include_toggle", True)

    wb = Workbook(); ws = wb.active
    ws.title = L("sheet_name", cfg); ws.sheet_view.showGridLines = False

    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.error = "Use YES or NO"; ws.add_data_validation(dv)

    sups = all_suppliers(df)
    rows = all_items_with_subitems(df)
    if not rows: print("ERROR: No items found"); sys.exit(1)

    # Column layout — image column always present in package_detail
    COL_INDENT = 2
    COL_BRAND  = 3
    COL_CODE   = 4
    COL_DESC   = 5
    COL_IMG    = 6
    if show_toggle:
        COL_INC = 7; COL_QTY = 8; SUP_START = 9
    else:
        COL_INC = None; COL_QTY = 7; SUP_START = 8
    last_col   = SUP_START + len(sups) - 1
    MK_START   = last_col + 3

    cr = 1

    # Title
    ws.row_dimensions[cr].height = 40
    ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=last_col)
    tc = ws.cell(row=cr, column=2, value=title)
    tc.font = Font(name='Segoe UI', bold=True, size=14, color=WH)
    tc.fill = HEADER_BLUE
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    pct_cells = write_markup_header(ws, cr, MK_START, sups, cfg)
    cr += 2

    # Winner calculation — sum items + subitems per supplier
    pkg_winner = ""
    pkg_min = float('inf')
    for s in sups:
        mkp = get_markup_pct(cfg, s) / 100.0
        total = 0
        for row in rows:
            p = price_for_any_type(df, s, row["code"], row["description"])
            if p > 0:
                total += p * (1 + mkp) if not is_service(row["description"]) else p
        if total > 0 and total * (1 + tr) < pkg_min:
            pkg_min = total * (1 + tr); pkg_winner = s

    # Include DV for package_detail (only if toggle enabled)
    if show_toggle:
        inc_dv_pd = DataValidation(type="list", formula1='"✓,✗"', allow_blank=False)
        ws.add_data_validation(inc_dv_pd)

    # Column headers
    ws.row_dimensions[cr].height = 28
    _incl_pd = L("incl_in_total", cfg)
    hdr_labels = ["", L("brand", cfg), L("code", cfg), L("description", cfg), L("image", cfg)] + ([_incl_pd] if show_toggle else []) + [L("qty", cfg)] + sups
    for i, h in enumerate(hdr_labels):
        ci = i + 2
        c = ws.cell(row=cr, column=ci, value=h if h in ("✓", _incl_pd) else h.upper())
        c.font = Font(name='Segoe UI', size=9, bold=True, color=TS)
        c.fill = WINNER_BG if h == pkg_winner else COL_HDR_BG
        c.alignment = Alignment(horizontal="center" if ci >= SUP_START else "left", vertical="center", wrap_text=True)
        c.border = COL_HDR_BD
    cr += 1

    sdr = cr
    ql = get_column_letter(COL_QTY)

    item_counter = 0
    for row in rows:
        is_sub = row["type"] == "subitem"
        r = cr

        if is_sub:
            row_h = 22
            base_fill = SUBITEM_ALT_BG if item_counter % 2 == 0 else SUBITEM_BG
            win_fill  = SUBITEM_WIN_BG
            font_size = 9
            font_color = "555555"
            indent_val = "↳"
            desc_indent = 3
        else:
            row_h = 30
            base_fill = ALT_ROW_BG if item_counter % 2 == 1 else WHITE_BG
            win_fill  = WINNER_BG
            font_size = 10
            font_color = TP
            indent_val = ""
            desc_indent = 1
            item_counter += 1

        ws.row_dimensions[r].height = row_h

        # Indent indicator
        ic = ws.cell(row=r, column=COL_INDENT, value=indent_val)
        ic.font = Font(name='Segoe UI', size=8, color="888888")
        ic.fill = base_fill
        ic.alignment = Alignment(horizontal="center", vertical="center")

        # Brand (blank for subitems)
        bc = ws.cell(row=r, column=COL_BRAND, value="" if is_sub else row["brand"])
        bc.font = Font(name='Segoe UI', size=font_size, color=font_color)
        bc.fill = base_fill; bc.border = THIN_BD

        # Code (blank for subitems)
        cc = ws.cell(row=r, column=COL_CODE, value="" if is_sub else row["code"])
        cc.font = Font(name='Segoe UI', size=font_size, color=font_color)
        cc.fill = base_fill; cc.border = THIN_BD

        # Description
        dc = ws.cell(row=r, column=COL_DESC, value=row["description"])
        dc.font = Font(name='Segoe UI', size=font_size, color=font_color,
                       italic=is_sub)
        dc.alignment = Alignment(vertical="center", horizontal="left", indent=desc_indent)
        dc.fill = base_fill; dc.border = THIN_BD

        # Image cell — always present, blank white for items, empty for subitems
        img_c = ws.cell(row=r, column=COL_IMG, value="")
        img_c.fill = WHITE_BG if not is_sub else base_fill
        img_c.border = THIN_BD

        # Include toggle (optional)
        if show_toggle:
            inc_c = ws.cell(row=r, column=COL_INC, value="✓")
            inc_c.font = Font(name='Segoe UI', size=10 if not is_sub else 9,
                              color="2C3E50", bold=True)
            inc_c.alignment = Alignment(horizontal="center", vertical="center")
            inc_c.fill = base_fill; inc_c.border = THIN_BD
            inc_dv_pd.add(inc_c)
        inc_cell = f"{get_column_letter(COL_INC)}{r}" if show_toggle else None

        # QTY
        qc = ws.cell(row=r, column=COL_QTY, value=1)
        qc.font = Font(name='Segoe UI', size=font_size, color=font_color, bold=not is_sub)
        qc.alignment = Alignment(horizontal="center", vertical="center")
        qc.fill = base_fill; qc.border = THIN_BD

        # Markup zone row
        prices = {s: price_for_any_type(df, s, row["code"], row["description"]) for s in sups}
        refs = write_markup_row(ws, r, MK_START, sups, pct_cells, row["description"], prices, dv)

        # Supplier price cells
        qty_cell = f"{ql}{r}"
        for si, s in enumerate(sups):
            col = SUP_START + si
            p = prices[s]
            c = ws.cell(row=r, column=col)
            if p > 0:
                c.value = markup_formula(qty_cell, refs[s])
                c.number_format = '$#,##0.00'
                c.font = Font(name='Segoe UI', size=font_size, color=font_color, italic=is_sub)
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = win_fill if s == pkg_winner else base_fill
            else:
                c.value = "—" if is_sub else ""
                c.font = Font(name='Segoe UI', size=font_size, color="CCCCCC")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = win_fill if s == pkg_winner else base_fill
            c.border = THIN_BD

        cr += 1

    # Apply red font CF for ✗ rows (only if toggle enabled)
    if show_toggle:
        apply_exclude_cf(ws, get_column_letter(COL_INC), sdr, cr - 1)

    # Subtotal
    ws.row_dimensions[cr].height = 32
    ws.merge_cells(start_row=cr, start_column=COL_INDENT, end_row=cr, end_column=COL_QTY)
    lc = ws.cell(row=cr, column=COL_INDENT, value=L("subtotal", cfg))
    lc.font = Font(name='Segoe UI', size=11, bold=True, color=TP)
    lc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    lc.border = Border(top=Side(style='medium', color="E5E5E5"))
    for si, s in enumerate(sups):
        col = SUP_START + si; cl = get_column_letter(col)
        if show_toggle:
            inc_cl = get_column_letter(COL_INC)
            formula = f'=SUMIF({inc_cl}{sdr}:{inc_cl}{cr-1},"✓",{cl}{sdr}:{cl}{cr-1})'
        else:
            formula = f'=SUM({cl}{sdr}:{cl}{cr-1})'
        c = ws.cell(row=cr, column=col, value=formula)
        c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(name='Segoe UI', size=11, bold=True, color=TP)
        c.border = Border(top=Side(style='medium', color="E5E5E5"))
        if s == pkg_winner: c.fill = WINNER_BG
    sub_row = cr; cr += 1

    # Tax
    ws.row_dimensions[cr].height = 28
    ws.merge_cells(start_row=cr, start_column=COL_INDENT, end_row=cr, end_column=COL_QTY)
    tc = ws.cell(row=cr, column=COL_INDENT, value=f"{L('tax', cfg)} {tl}")
    tc.font = Font(name='Segoe UI', size=10, color=TS)
    tc.alignment = Alignment(vertical="center", horizontal="right", indent=1); tc.border = SUBTLE_BD
    for si, s in enumerate(sups):
        col = SUP_START + si; cl = get_column_letter(col)
        c = ws.cell(row=cr, column=col, value=f"={cl}{sub_row}*{tr}")
        c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(name='Segoe UI', size=10, color=TS); c.border = SUBTLE_BD
        if s == pkg_winner: c.fill = WINNER_BG
    tax_r = cr; cr += 1

    # Total
    ws.row_dimensions[cr].height = 40
    ws.merge_cells(start_row=cr, start_column=COL_INDENT, end_row=cr, end_column=COL_QTY)
    tl_cell = ws.cell(row=cr, column=COL_INDENT, value=L("total", cfg))
    tl_cell.font = Font(name='Segoe UI', size=13, bold=True, color=TP)
    tl_cell.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    tl_cell.border = TOTAL_BD
    for si, s in enumerate(sups):
        col = SUP_START + si; cl = get_column_letter(col)
        c = ws.cell(row=cr, column=col, value=f"={cl}{sub_row}+{cl}{tax_r}")
        c.font = Font(name='Segoe UI', bold=True, size=13, color=TP)
        c.number_format = '$#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = TOTAL_BD
        if s == pkg_winner: c.fill = WINNER_BG
    cr += 1

    # Winner row
    ws.row_dimensions[cr].height = 28
    ws.merge_cells(start_row=cr, start_column=COL_INDENT, end_row=cr, end_column=COL_QTY)
    ws.cell(row=cr, column=COL_INDENT, value="").fill = WINNER_BG
    for si, s in enumerate(sups):
        col = SUP_START + si; c = ws.cell(row=cr, column=col)
        if s == pkg_winner:
            c.value = L("best_price", cfg)
            c.font = Font(name='Segoe UI', bold=True, size=11, color="1B5E20")
        else:
            c.value = ""
        c.fill = WINNER_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style='thin', color="E5E5E5"))

    # Column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions[get_column_letter(COL_INDENT)].width = 4
    ws.column_dimensions[get_column_letter(COL_BRAND)].width = 14
    ws.column_dimensions[get_column_letter(COL_CODE)].width = 16
    ws.column_dimensions[get_column_letter(COL_DESC)].width = 34
    ws.column_dimensions[get_column_letter(COL_IMG)].width = 12  # IMAGE always present
    if show_toggle:
        ws.column_dimensions[get_column_letter(COL_INC)].width = 14
    ws.column_dimensions[get_column_letter(COL_QTY)].width = 7
    for i in range(len(sups)):
        ws.column_dimensions[get_column_letter(SUP_START + i)].width = 16
    ws.column_dimensions[get_column_letter(MK_START)].width = 22
    for si in range(len(sups)):
        ws.column_dimensions[get_column_letter(MK_START + 1 + si*2)].width = 14
        ws.column_dimensions[get_column_letter(MK_START + 2 + si*2)].width = 8

    wb.save(cfg["output_xlsx"])
    item_count = sum(1 for r in rows if r["type"] == "item")
    sub_count  = sum(1 for r in rows if r["type"] == "subitem")
    print(f"[package_detail] Saved: {cfg['output_xlsx']} | {item_count} items + {sub_count} subitems x {len(sups)} suppliers | Tax: {tl}")
    mk = cfg.get("markup", {})
    if any(v > 0 for v in mk.values()): print(f"Markup: {mk}")


# ===================== MAIN =====================

def main():
    ap = argparse.ArgumentParser(description="FF&E Price Analysis Excel Generator")
    ap.add_argument("--config", "-c", required=True)
    args = ap.parse_args()

    if not os.path.exists(args.config): print(f"ERROR: Config not found: {args.config}"); sys.exit(1)
    with open(args.config) as f: cfg = json.load(f)

    for fld in ("input_csv", "output_xlsx"):
        if fld not in cfg: print(f"ERROR: '{fld}' required"); sys.exit(1)

    if "province" in cfg:
        p = cfg["province"].upper().strip()
        if p not in PROVINCE_TAX: print(f"ERROR: Unknown province '{p}'"); sys.exit(1)
        cfg["tax_rate"] = PROVINCE_TAX[p]
        cfg["tax_label"] = f"{cfg['tax_rate']}% {p}"
    elif "tax_rate" in cfg:
        cfg["tax_label"] = f"{cfg['tax_rate']}%"
    else: print("ERROR: Provide 'tax_rate' or 'province'"); sys.exit(1)

    cfg.setdefault("lang", "en")
    cfg.setdefault("mode", "individual")
    cfg.setdefault("include_specs", False)
    cfg.setdefault("include_image", False)
    cfg.setdefault("include_toggle", True)
    cfg.setdefault("markup", {})
    # Title default uses localized fallback — only set if not already provided
    if "title" not in cfg:
        cfg["title"] = L("default_title_ind", cfg) if cfg["mode"] == "individual" else L("default_title_pkg", cfg)

    if not os.path.exists(cfg["input_csv"]): print(f"ERROR: CSV not found: {cfg['input_csv']}"); sys.exit(1)

    df = load_csv(cfg["input_csv"])
    if (USE_PD and df.empty) or (not USE_PD and not df): print("ERROR: CSV empty"); sys.exit(1)

    if cfg["mode"] == "package_detail": build_package_detail(df, cfg)
    elif cfg["mode"] == "package": build_package(df, cfg)  # legacy, not recommended
    else: build_individual(df, cfg)

if __name__ == "__main__":
    main()
