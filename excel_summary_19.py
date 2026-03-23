import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, numbers, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io



st.set_page_config(layout="wide")

# -------------------------------------------------
# TRANSLATIONS
# -------------------------------------------------
TRANSLATIONS = {
    "en": {
        "app_title": "CSV to Excel Price Analysis",
        "upload_csv": "Upload CSV",
        "csv_loaded": "CSV loaded",
        "review_table": "Review Source Table",
        "tax_settings": "Tax Settings",
        "tax_percent": "Tax Percentage",
        "preview_title": "Price Analysis Preview (HTML Table)",
        "generate_excel_title": "Generate Final Excel",
        "generate_excel_btn": "Generate Excel File",
        "download_excel": "Download Excel",
        "upload_prompt": "Upload or generate data to see the price analysis preview.",
        "pricing_mode_label": "Pricing Mode",
        "individual": "Individual",
        "package": "Package",
        "details": "DETAILS",
        "image": "IMAGE",
        "qty": "QTY",
        "line_item": "LINE ITEM",
        "photo": "[ PHOTO ]",
        "brand": "BRAND",
        "code": "CODE",
        "power": "POWER",
        "package_price": "Package Price",
        "total_before_tax": "Total Before Tax",
        "tax_label": "Tax",
        "total": "Total",
        "specs_title": "SPECS & DESCRIPTION",
        "specs_placeholder": "Enter item specifications, dimensions, and technical details here...",
        "option": "Option",
        "price_analysis_sheet": "Price Analysis",
    },
    "fr": {
        "app_title": "Analyse de Prix CSV vers Excel",
        "upload_csv": "Telecharger CSV",
        "csv_loaded": "CSV charge",
        "review_table": "Reviser le tableau source",
        "tax_settings": "Parametres de taxes",
        "tax_percent": "Pourcentage de taxe",
        "preview_title": "Apercu de l'analyse de prix (Tableau HTML)",
        "generate_excel_title": "Generer le fichier Excel final",
        "generate_excel_btn": "Generer le fichier Excel",
        "download_excel": "Telecharger Excel",
        "upload_prompt": "Telechargez ou generez des donnees pour voir l'apercu.",
        "pricing_mode_label": "Mode de tarification",
        "individual": "Individuel",
        "package": "Forfait",
        "details": "DETAILS",
        "image": "IMAGE",
        "qty": "QTE",
        "line_item": "ARTICLE",
        "photo": "[ PHOTO ]",
        "brand": "MARQUE",
        "code": "CODE",
        "power": "ALIMENTATION",
        "package_price": "Prix forfaitaire",
        "total_before_tax": "Total avant taxes",
        "tax_label": "Taxe",
        "total": "Total",
        "specs_title": "SPECIFICATIONS ET DESCRIPTION",
        "specs_placeholder": "Entrez les specifications, dimensions et details techniques ici...",
        "option": "Option",
        "price_analysis_sheet": "Analyse de Prix",
    },
}

# -------------------------------------------------
# SESSION STATE
# -------------------------------------------------
if "df" not in st.session_state:
    st.session_state.df = None

# -------------------------------------------------
# SIDEBAR CONFIG
# -------------------------------------------------
with st.sidebar:
    lang = st.selectbox("Language / Langue", options=["en", "fr"], format_func=lambda x: "English" if x == "en" else "Francais")
    T = TRANSLATIONS[lang]

    pricing_mode = st.radio(
        T["pricing_mode_label"],
        options=["individual", "package"],
        format_func=lambda x: T["individual"] if x == "individual" else T["package"]
    )

T = TRANSLATIONS[lang]

# -------------------------------------------------
# HEADER
# -------------------------------------------------
st.title(T["app_title"])

# -------------------------------------------------
# UPLOAD CSV
# -------------------------------------------------
uploaded_file = st.file_uploader(
    T["upload_csv"],
    type=["csv"]
)

if uploaded_file:
    df = pd.read_csv(uploaded_file)

    for col in ["type", "supplier", "brand", "code", "description", "Power Type"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    st.session_state.df = df.copy()
    st.success(T["csv_loaded"])



# -------------------------------------------------
# EDIT SOURCE TABLE
# -------------------------------------------------
if st.session_state.df is not None:
    st.subheader(T["review_table"])
    st.session_state.df = st.data_editor(
        st.session_state.df,
        use_container_width=True,
        num_rows="dynamic"
    )

# -------------------------------------------------
# TAX INPUT
# -------------------------------------------------
st.subheader(T["tax_settings"])
tax_percent = st.number_input(T["tax_percent"], min_value=0.0, value=12.0)

# -------------------------------------------------
# HTML PREVIEW (EXCEL-STYLE)
# -------------------------------------------------
st.subheader(T["preview_title"])

def generate_html_table(df, tax_percent, T, pricing_mode="individual"):
    tax_rate = tax_percent / 100

    html = """
    <div style="overflow-x:auto;">
    <style>
        table {
            border-collapse: collapse !important;
            width: 100%;
            margin-bottom: 40px;
            font-family: Arial, sans-serif;
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #bfbfbf;
        }

        th, td {
            border: 1px solid #bfbfbf !important;
            padding: 6px 8px;
            vertical-align: middle;
            text-align: left;
            background-clip: padding-box;
        }

        th {
            background-color: #dae9f8;
            font-weight: 600;
        }

        .total-row td {
            background-color: #fce4d6;
            font-weight: bold;
        }
    </style>
    """

    main_items = df[
        (df["type"] == "item") &
        df["Power Type"].notna() &
        (df["Power Type"] != "")
    ]

    for code, power_type in main_items[["code", "Power Type"]].drop_duplicates().values:

        items_for_code = df[
            (df["code"] == code) &
            (
                (df["Power Type"] == power_type) |
                (df["Power Type"].isna()) |
                (df["Power Type"] == "")
            ) &
            (df["type"].isin(["item", "subitem"]))
        ]

        suppliers = items_for_code["supplier"].unique()
        brand = items_for_code[items_for_code["type"] == "item"].iloc[0]["brand"]
        descriptions = items_for_code["description"].unique()

        if pricing_mode == "package":
            # Package: descriptions + package price row + tax + total
            body_rows = len(descriptions) + 3
        else:
            # Individual: descriptions + tax + total
            body_rows = len(descriptions) + 2

        html += "<table>"

        # HEADER
        html += "<tr>"
        html += f"<th>{T['details']}</th><th></th><th>{T['qty']}</th><th>{T['line_item']}</th>"
        for s in suppliers:
            html += f"<th>{s}</th>"
        html += "</tr>"

        # Compute totals (sum of all items per supplier) for both modes
        totals = {}
        for s in suppliers:
            s_total = 0
            for desc in descriptions:
                row = items_for_code[
                    (items_for_code["supplier"] == s) &
                    (items_for_code["description"] == desc)
                ]
                s_total += float(row["price"].iloc[0]) if not row.empty else 0
            totals[s] = s_total

        # FIRST ITEM ROW (with DETAILS merged cell)
        first_desc = descriptions[0]

        html += "<tr>"
        html += f"""
            <td rowspan="{body_rows}">
                <b>{T['brand']}</b><br>{brand}<br><br>
                <b>{T['code']}</b><br>{code}<br><br>
                <b>{T['power']}</b><br>{power_type}
            </td>
            <td rowspan="{body_rows}"></td>
            <td>1</td>
            <td>{first_desc}</td>
        """

        if pricing_mode == "individual":
            for s in suppliers:
                row = items_for_code[
                    (items_for_code["supplier"] == s) &
                    (items_for_code["description"] == first_desc)
                ]
                price = float(row["price"].iloc[0]) if not row.empty else 0
                html += f"<td>${price:,.2f}</td>"
        else:
            for _ in suppliers:
                html += "<td></td>"

        html += "</tr>"

        # REMAINING ITEM ROWS
        for desc in descriptions[1:]:
            html += "<tr>"
            html += f"<td>1</td><td>{desc}</td>"

            if pricing_mode == "individual":
                for s in suppliers:
                    row = items_for_code[
                        (items_for_code["supplier"] == s) &
                        (items_for_code["description"] == desc)
                    ]
                    price = float(row["price"].iloc[0]) if not row.empty else 0
                    html += f"<td>${price:,.2f}</td>"
            else:
                for _ in suppliers:
                    html += "<td></td>"

            html += "</tr>"

        # PACKAGE PRICE ROW (package mode only)
        if pricing_mode == "package":
            html += "<tr>"
            html += f"<td></td><td><b>{T['package_price']}</b></td>"
            for s in suppliers:
                html += f"<td>${totals[s]:,.2f}</td>"
            html += "</tr>"

        # TAX ROW
        html += "<tr>"
        html += f"<td></td><td><b>{T['tax_label']}</b></td>"
        for _ in suppliers:
            html += f"<td>{tax_percent:.2f}%</td>"
        html += "</tr>"

        # TOTAL ROW
        html += "<tr class='total-row'>"
        html += f"<td></td><td>{T['total']}</td>"
        for s in suppliers:
            total = totals[s] * (1 + tax_rate)
            html += f"<td>${total:,.2f}</td>"
        html += "</tr>"

        html += "</table>"

    html += "</div>"
    return html


# RENDER HTML (LIVE, REACTIVE)
if (
    "df" in st.session_state
    and st.session_state.df is not None
    and not st.session_state.df.empty
):
    html = generate_html_table(st.session_state.df, tax_percent, T, pricing_mode)
    st.markdown(html, unsafe_allow_html=True)
else:
    st.info(T["upload_prompt"])

# -------------------------------------------------
# GENERATE FINAL EXCEL (MINIMALIST FORMATTING - RFQ STYLE)
# -------------------------------------------------
st.subheader(T["generate_excel_title"])

if st.button(T["generate_excel_btn"]):
    df = st.session_state.df
    tax_rate = tax_percent / 100

    wb = Workbook()
    ws = wb.active
    ws.title = T["price_analysis_sheet"]
    ws.sheet_view.showGridLines = False

    # --- DESIGN TOKENS ---
    HEADER_BLUE = PatternFill(start_color="288AD6", end_color="288AD6", fill_type="solid")
    DETAILS_BG = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    COLUMN_HEADER_BG = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    WINNER_BG = PatternFill(start_color="F2FAF2", end_color="F2FAF2", fill_type="solid")
    SPECS_BG = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

    SUBTLE_BORDER = Border(bottom=Side(style='thin', color="F0F0F0"))
    COLUMN_HEADER_BORDER = Border(bottom=Side(style='medium', color="E5E5E5"))
    TOTAL_BORDER = Border(top=Side(style='medium', color="288AD6"))

    TEXT_PRIMARY = "1D1D1F"
    TEXT_SECONDARY = "86868B"
    WHITE = "FFFFFF"

    current_row = 1

    main_items = df[
        (df["type"] == "item") &
        df["Power Type"].notna() &
        (df["Power Type"] != "")
    ]

    for opt_idx, (code, power_type) in enumerate(
        main_items[["code", "Power Type"]].drop_duplicates().values, 1
    ):
        items_for_code = df[
            (df["code"] == code) &
            (
                (df["Power Type"] == power_type) |
                (df["Power Type"].isna()) |
                (df["Power Type"] == "")
            ) &
            (df["type"].isin(["item", "subitem"]))
        ]

        suppliers = list(items_for_code["supplier"].unique())
        brand = items_for_code[items_for_code["type"] == "item"].iloc[0]["brand"]
        descriptions = list(items_for_code["description"].unique())

        # Compute package totals per supplier (sum of all line items)
        package_totals = {}
        for sup in suppliers:
            sup_items = items_for_code[items_for_code["supplier"] == sup]
            package_totals[sup] = float(sup_items["price"].sum())

        # --- DETERMINE WINNER (LOWEST TOTAL PRICE) ---
        winner_supplier = min(package_totals, key=package_totals.get)

        # === 1. OPTION TITLE (BLUE HEADER) ===
        title_row = current_row
        ws.row_dimensions[title_row].height = 40

        last_col = 6 + len(suppliers) - 1
        ws.merge_cells(
            start_row=title_row,
            start_column=2,
            end_row=title_row,
            end_column=last_col
        )

        title_cell = ws.cell(row=title_row, column=2, value=f"{T['option']} {opt_idx:02d}")
        title_cell.font = Font(name='Segoe UI', bold=True, size=14, color=WHITE)
        title_cell.fill = HEADER_BLUE
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

        # === 2. COLUMN HEADERS ===
        header_row = title_row + 1
        ws.row_dimensions[header_row].height = 28

        headers = [T["details"], T["image"], T["qty"], T["line_item"]] + suppliers
        for i, h in enumerate(headers):
            col_idx = i + 2
            cell = ws.cell(row=header_row, column=col_idx, value=h.upper())
            cell.font = Font(name='Segoe UI', size=9, bold=True, color=TEXT_SECONDARY)
            cell.fill = WINNER_BG if h == winner_supplier else COLUMN_HEADER_BG
            cell.alignment = Alignment(
                horizontal="center" if col_idx >= 6 else "left",
                vertical="center"
            )
            cell.border = COLUMN_HEADER_BORDER

        # === 3. CONTENT ROWS ===
        start_data_row = header_row + 1
        num_item_rows = len(descriptions)

        # DETAILS column (merged vertically through tax row)
        detail_val = f"{T['brand']}\n{brand}\n\n{T['code']}\n{code}\n\n{T['power']}\n{power_type}"
        d_cell = ws.cell(row=start_data_row, column=2, value=detail_val)
        d_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=1)
        d_cell.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
        d_cell.fill = DETAILS_BG

        # Merge depth: item rows + package price row (if package) + tax row
        merge_depth = num_item_rows + (1 if pricing_mode == "package" else 0) + 1
        ws.merge_cells(
            start_row=start_data_row, start_column=2,
            end_row=start_data_row + merge_depth, end_column=2
        )

        # IMAGE placeholder (merged same depth)
        img_cell = ws.cell(row=start_data_row, column=3, value=T["photo"])
        img_cell.alignment = Alignment(horizontal="center", vertical="center")
        img_cell.font = Font(name='Segoe UI', size=10, color="CCCCCC", italic=True)
        img_cell.fill = DETAILS_BG
        ws.merge_cells(
            start_row=start_data_row, start_column=3,
            end_row=start_data_row + merge_depth, end_column=3
        )

        # ITEM ROWS
        for idx, desc in enumerate(descriptions):
            r_num = start_data_row + idx
            ws.row_dimensions[r_num].height = 32

            qty_cell = ws.cell(row=r_num, column=4, value=1)
            qty_cell.alignment = Alignment(horizontal="center", vertical="center")
            qty_cell.font = Font(name='Segoe UI', size=11, color=TEXT_PRIMARY, bold=True)
            qty_cell.border = SUBTLE_BORDER

            desc_cell = ws.cell(row=r_num, column=5, value=desc)
            desc_cell.font = Font(name='Segoe UI', size=11, color=TEXT_PRIMARY)
            desc_cell.alignment = Alignment(vertical="center", horizontal="left")
            desc_cell.border = SUBTLE_BORDER

            qty_letter = get_column_letter(4)
            for s_idx, sup in enumerate(suppliers):
                col = 6 + s_idx
                price_row = items_for_code[
                    (items_for_code["supplier"] == sup) &
                    (items_for_code["description"] == desc)
                ]
                price = float(price_row["price"].iloc[0]) if not price_row.empty else 0

                if pricing_mode == "individual":
                    cell = ws.cell(row=r_num, column=col, value=f"={qty_letter}{r_num}*{price}")
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Segoe UI', size=11, color=TEXT_PRIMARY)
                else:
                    # Package mode: leave price cells empty
                    cell = ws.cell(row=r_num, column=col, value="")
                    cell.font = Font(name='Segoe UI', size=11, color=TEXT_SECONDARY)

                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = SUBTLE_BORDER
                if sup == winner_supplier:
                    cell.fill = WINNER_BG

        # Current row pointer after item rows
        next_row = start_data_row + num_item_rows

        # === 4a. PACKAGE PRICE ROW (package mode only) ===
        if pricing_mode == "package":
            pkg_row = next_row
            ws.row_dimensions[pkg_row].height = 32

            ws.merge_cells(
                start_row=pkg_row, start_column=4,
                end_row=pkg_row, end_column=5
            )
            pkg_label = ws.cell(row=pkg_row, column=4, value=T["package_price"])
            pkg_label.font = Font(name='Segoe UI', size=11, bold=True, color=TEXT_PRIMARY)
            pkg_label.alignment = Alignment(vertical="center", horizontal="left")
            pkg_label.border = SUBTLE_BORDER

            for s_idx, sup in enumerate(suppliers):
                col = 6 + s_idx
                pkg_cell = ws.cell(row=pkg_row, column=col, value=package_totals[sup])
                pkg_cell.number_format = '$#,##0.00'
                pkg_cell.alignment = Alignment(horizontal="right", vertical="center")
                pkg_cell.font = Font(name='Segoe UI', size=11, bold=True, color=TEXT_PRIMARY)
                pkg_cell.border = SUBTLE_BORDER
                if sup == winner_supplier:
                    pkg_cell.fill = WINNER_BG

            next_row += 1

        # === 4b. TOTAL BEFORE TAX ROW (individual mode only) ===
        if pricing_mode == "individual":
            total_before_tax_row = next_row
            ws.row_dimensions[total_before_tax_row].height = 32

            ws.merge_cells(
                start_row=total_before_tax_row, start_column=4,
                end_row=total_before_tax_row, end_column=5
            )
            tbt_label = ws.cell(row=total_before_tax_row, column=4, value=T["total_before_tax"])
            tbt_label.font = Font(name='Segoe UI', size=11, bold=True, color=TEXT_PRIMARY)
            tbt_label.alignment = Alignment(vertical="center", horizontal="left")
            tbt_label.border = SUBTLE_BORDER

            for s_idx, sup in enumerate(suppliers):
                col = 6 + s_idx
                col_letter = get_column_letter(col)
                tbt_cell = ws.cell(
                    row=total_before_tax_row,
                    column=col,
                    value=f"=SUM({col_letter}{start_data_row}:{col_letter}{total_before_tax_row-1})"
                )
                tbt_cell.number_format = '$#,##0.00'
                tbt_cell.alignment = Alignment(horizontal="right", vertical="center")
                tbt_cell.font = Font(name='Segoe UI', size=11, bold=True, color=TEXT_PRIMARY)
                tbt_cell.border = SUBTLE_BORDER
                if sup == winner_supplier:
                    tbt_cell.fill = WINNER_BG

            subtotal_row = total_before_tax_row
            next_row += 1
        else:
            subtotal_row = pkg_row

        # === 5. TAX ROW ===
        tax_row = next_row
        ws.row_dimensions[tax_row].height = 28

        ws.merge_cells(
            start_row=tax_row, start_column=4,
            end_row=tax_row, end_column=5
        )
        tax_label_cell = ws.cell(row=tax_row, column=4, value=f"{T['tax_label']} ({int(tax_rate*100)}%)")
        tax_label_cell.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
        tax_label_cell.alignment = Alignment(vertical="center", horizontal="left")
        tax_label_cell.border = SUBTLE_BORDER

        for s_idx, sup in enumerate(suppliers):
            col = 6 + s_idx
            col_letter = get_column_letter(col)
            t_cell = ws.cell(
                row=tax_row, column=col,
                value=f"={col_letter}{subtotal_row}*{tax_rate}"
            )
            t_cell.number_format = '$#,##0.00'
            t_cell.alignment = Alignment(horizontal="right", vertical="center")
            t_cell.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
            t_cell.border = SUBTLE_BORDER
            if sup == winner_supplier:
                t_cell.fill = WINNER_BG

        # === 6. FINAL TOTAL ROW ===
        total_row = tax_row + 1
        ws.row_dimensions[total_row].height = 40

        ws.merge_cells(
            start_row=total_row, start_column=2,
            end_row=total_row, end_column=5
        )
        empty_cell = ws.cell(row=total_row, column=2, value="")
        empty_cell.fill = DETAILS_BG
        empty_cell.border = TOTAL_BORDER

        for s_idx, sup in enumerate(suppliers):
            col = 6 + s_idx
            col_letter = get_column_letter(col)
            tot_cell = ws.cell(
                row=total_row, column=col,
                value=f"={col_letter}{subtotal_row}+{col_letter}{tax_row}"
            )
            tot_cell.font = Font(name='Segoe UI', bold=True, size=13, color=TEXT_PRIMARY)
            tot_cell.number_format = '$#,##0.00'
            tot_cell.alignment = Alignment(horizontal="right", vertical="center")
            tot_cell.border = TOTAL_BORDER
            if sup == winner_supplier:
                tot_cell.fill = WINNER_BG

        # === 7. SPECS & DESCRIPTION BLOCK ===
        specs_row = total_row + 1
        ws.row_dimensions[specs_row].height = 60

        ws.merge_cells(
            start_row=specs_row, start_column=2,
            end_row=specs_row, end_column=last_col
        )
        specs_content = ws.cell(
            row=specs_row, column=2,
            value=f"{T['specs_title']}\n\n{T['specs_placeholder']}"
        )
        specs_content.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
        specs_content.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="left", indent=2
        )
        specs_content.fill = SPECS_BG
        specs_content.border = Border(bottom=Side(style='thin', color="F0F0F0"))

        current_row = specs_row + 3

    # === COLUMN WIDTH ADJUSTMENTS ===
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 35

    for i in range(len(suppliers)):
        ws.column_dimensions[get_column_letter(6+i)].width = 16

    output = io.BytesIO()
    wb.save(output)

    st.download_button(
        T["download_excel"],
        data=output.getvalue(),
        file_name=f"price_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
